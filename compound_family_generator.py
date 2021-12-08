# add timer
# meaning = meaning OR buddhadatta
# exlcude those with no meaning, onlybuddhdatta?

import csv
import openpyxl as xl
import re
from xlsxwriter.workbook import Workbook
import logging

#setup logger
logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%I:%M:%S')

# convert csv to excel

convert = input (f"convert csv to excel (y/n)? ")
if convert == "y":
    logging.warning(f"converting csv to excel sheet")
    csv_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary-full.csv"
    xlsx_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary-full.xlsx"
    workbook = Workbook(xlsx_file)
    worksheet = workbook.add_worksheet()
    csv_reader = csv.reader(open(csv_file,'rt'),delimiter="\t")

    for row, data in enumerate(csv_reader):
        worksheet.write_row(row, 0, data)
    workbook.close()
    
if convert == "n":
    pass

# open excel sheet
logging.warning(f"loading excel sheet")

wb = xl.load_workbook("Pāli English Dictionary-full.xlsx")
sheet = wb["Sheet1"]
last_row = sheet.max_row + 1
row_number = 2
compound_family_list = []

# open csv
logging.warning (f"opening text file")
txt_file = open ("compound_family_list.txt", 'w', encoding= "'utf-8")

# extract compound families
logging.warning (f"extracting compound family names")

compound_family_list = []

for row in range(row_number, last_row):
    family_cell = sheet.cell(row, 26)
    metadata = sheet.cell(row, 54)
    meaning = sheet.cell(row, 11)

    if (family_cell.value != None
    and metadata.value == None
    and meaning.value != None):
        compound_family_list += family_cell.value.split()

compound_family_list.sort()
compound_family_list = list(dict.fromkeys(compound_family_list))

compound_family_count = 0

logging.warning (f"writing compound_family_list.txt")
for family in compound_family_list:
    compound_family_count += 1
    family_value = str(family)
    txt_file.write (f"{compound_family_count}. {family_value}\n")

# open csv
logging.warning (f"opening compound_families.csv")
txt_file = open ("compound_families.csv", 'w', encoding= "'utf-8")

# write families
logging.warning (f"writing families")

line_break = ("~" * 40)
print(line_break)

compound_family_count = 0

txt_file.write (f"headword\tpos\tmeaning\tconstruction\n")

for family in compound_family_list:
    family_value = str(family)
    compound_family_count += 1
    print(f"{compound_family_count}. {family_value}")
    txt_file.write (f"\n")
    txt_file.write (f"{compound_family_count}. {family}\n")
    
    for row in range(2, last_row):
        compound_family_cell = sheet.cell(row, 26)
        compound_family_cell_value = str(compound_family_cell.value)
        meaning = sheet.cell(row, 11)
        headword = sheet.cell(row, 1)
        metadata = sheet.cell(row, 54)
        pos = sheet.cell(row, 4)
        
        construction = sheet.cell(row, 27)
        if construction.value == None:
            construction_value = str(construction.value)
            pattern = "None"
            replace = ""
            string = construction_value
            construction_value_new = re.sub(pattern ,replace, string)
        else:
            construction_value = str(construction.value)
            pattern = "<br.+$"
            replace = ""
            string = construction_value
            construction_value_new = re.sub(pattern ,replace, string)

        search_pattern = re.compile(rf"\b{family}\b")
        test = re.search(search_pattern, compound_family_cell_value)
                
        if (test != None
        and meaning.value != None
        and metadata.value != "yes"):
            txt_file.write (f"{headword.value}\t{pos.value}\t{meaning.value}\t{construction_value_new}\n")
        else:
            continue

logging.warning(f"{compound_family_count} familes")
logging.warning(f"saved to {txt_file.name}")
logging.warning(f"fin")


# 1	Pāli1
# 2	Pāli2
# 3	Fin
# 4	POS
# 5	Grammar
# 6	Derived from
# 7	Neg
# 8	Verb
# 9	Trans
# 10	Case
# 11	Meaning IN CONTEXT
# 12	Literal Meaning
# 13	Non IA
# 14	Sanskrit
# 15	Sk Root
# 16	Sk Root Mn
# 17	Cl
# 18	Pāli Root
# 19	Root In Comps
# 20	V
# 21	Grp
# 22	Sgn
# 23	Root Meaning
# 24	Base
# 25	Family
# 26	Family2
# 27	Construction
# 28	Derivative
# 29	Suffix
# 30	Phonetic Changes
# 31	Compound
# 32	Compound Construction
# 33	Non-Root In Comps
# 34	Source1
# 35	Sutta1
# 36	Example1
# 37	Source 2
# 38	Sutta2
# 39	Example 2
# 40	Antonyms
# 41	Synonyms – different word
# 42	Variant – same constr or diff reading
# 43	Commentary
# 44	Notes
# 45	Cognate
# 46	Category
# 47	Link
# 48	Stem
# 49	Pattern
# 50	Buddhadatta
# 51	22
# 52	Pāli1 ≠ const
# 53	test dupl
# 54	Metadata